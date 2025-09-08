import openpyxl
import os
from openpyxl.styles import Font, PatternFill
from calculations import calcular_puntaje_total


#Exporta los resultados de estudiantes a un archivo excel
# records (list[dict]): Lista de registros de estudiantes con sus datos.

def export_excel(records: list[dict], filename: str = "puntajes_export.xlsx") -> None:
    wb = openpyxl.Workbook()

    # Datos de Resultados de puntajes finales
    ws_resultado = wb.active
    ws_resultado.title = "Resultados"

    headers_resultados = [
        "Rut", "Nombre", "Curso",
        "Elección 1", "Elección 2", "Elección 3", "Elección 4", "Elección 5",
        "Anotaciones", "Notas Base", "Tecnología", "SIMCE", "Asistencia",
        "Entrevistas", "Puntaje Total"
    ]
    ws_resultado.append(headers_resultados)

    #Estilo encabezado resultado
    for cell in ws_resultado[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    #Hoja Detalles muestra informacion detallada del estudiante
    ws_detalle = wb.create_sheet("Detalles")

    headers_detalle = [
        "Rut", "Nombre", "Curso",
        "Anot. Leves 1", "Anot. Graves 1", "Anot. Gravísimas 1",
        "Anot. Leves 2", "Anot. Graves 2", "Anot. Gravísimas 2",
        "Prom. Lneguaje 1", "Prom. Lenguaje 2",
        "Prom. Matemática 1", "Prom. Matemática 2",
        "Prom. General 1", "Prom. General 2",
        "Prom.Tecno 1", "Prom.Tecno 2", "Punt. Prueba Tecno",
        "SIMCE Lenguaje 2", "SIMCE Matemática 2",
        "Asistencia 1", "Asistencia 2",
        "Puntaje Entrevistas"

    ]
    ws_detalle.append(headers_detalle)

    # Estilo encabezado detalle
    for cell in ws_detalle[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type='solid')


    #Rellenar datos
    for record in records:
        puntajes = calcular_puntaje_total(record, records)

        #fila de resultados
        ws_resultado.append([
            record.get("rut"), record.get("full_name"), record.get("course"),
            record.get("eleccion_1"), record.get("eleccion_2"), record.get("eleccion_3"),
            record.get("eleccion_4"), record.get("eleccion_5"),
            puntajes["anotaciones"], puntajes["notas_base"], puntajes["tecnologia"], puntajes["simce"], puntajes["asistencia"],
            puntajes["entrevistas"], puntajes["total_final_puntajes"]
        ])

        #fila de detalles
        ws_detalle.append([
            record.get("rut"),
            record.get("full_name"),
            record.get("course"),
            record.get("anotaciones_leves_1"),
            record.get("anotaciones_graves_1"),
            record.get("anotaciones_gravisimas_2"),
            record.get("anotaciones_leves_2"),
            record.get("anotaciones_graves_2"),
            record.get("anotaciones_gravisimas_2"),
            record.get("promedio_lenguaje_1"),
            record.get("promedio_lenguaje_2"),
            record.get("promedio_matematica_1"),
            record.get("promedio_matematica_2"),
            record.get("promedio_general_1"),
            record.get("promedio_general_2"),
            record.get("promedio_tecnologia_1"),
            record.get("promedio_tecnologia_2"),
            record.get("puntaje_prueba_tecnologia"),
            record.get("puntaje_simce_lenguaje_2"),
            record.get("puntaje_simce_matematica_2"),
            record.get("asistencia_1"),
            record.get("asistencia_2"),
            record.get("puntaje_entrevistas")
        ])

        #Ajuste ancho columnas
        for ws in [ws_resultado, ws_detalle]:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max_len + 2

        #verificar si el archivo ya existe y crear uno nuevo
        if os.path.exists(filename):
            base, ext = os.path.splitext(filename)
            i =1
            while os.path.exists(f"{base}_{i}{ext}"): i += 1
            filename = f"{base}_{i}{ext}"

    #Guardar Archivo
    wb.save(filename)
    print(f"✅ Archivo Excel generado: {filename}")







